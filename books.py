from download_page import download_page
from bs4 import BeautifulSoup
from openpyxl import Workbook
import time

DOWNLOAD_URL = "https://book.douban.com/top250"
SLEEP_TIME = 3


def get_li(doc):
    bs = BeautifulSoup(doc, 'html.parser')
    tables = bs.find('div', class_='indent').find_all('table')
    title_list = []
    detail_list = []
    quote_list = []
    star_list = []
    for t in tables:
        td = t.find_all('td', attrs={'valign': 'top'})[-1]
        title = td.find('div', class_='pl2').find('a')['title']
        title_list.append(title)
        detail = td.find('p', class_='pl').get_text()
        detail_list.append(detail)
        quote = td.find('span', class_='inq')
        if quote:
            quote_list.append(quote.get_text())
        else:
            quote_list.append("无")
        star = td.find('span', class_='rating_nums').get_text()
        star_list.append(star)
    next_page = bs.find('link', attrs={'rel': 'next'})
    if next_page:
        return title_list, detail_list, quote_list, star_list, next_page['href']
    else:
        return title_list, detail_list, quote_list, star_list, None


def main():
    print('开始获取数据...')
    url = DOWNLOAD_URL
    title_list = []
    detail_list = []
    quote_list = []
    star_list = []

    while url:
        doc = download_page(url)
        tl, dl, ql, sl, url = get_li(doc)
        title_list = title_list + tl
        detail_list = detail_list + dl
        quote_list = quote_list + ql
        star_list = star_list + sl
        print('已获取数据%s条' % len(title_list))
        time.sleep(SLEEP_TIME)

    wb = Workbook()
    ws = wb.active
    ws.title = '豆瓣书籍top250'
    print('数据获取结束...')
    print('开始写入excel...')

    for (t, d, q, s) in zip(title_list, detail_list, quote_list, star_list):
        index = title_list.index(t) + 1
        ws['A%s' % index] = t
        ws['B%s' % index] = s
        ws['C%s' % index] = d
        ws['D%s' % index] = q

    wb.save(filename='豆瓣书籍top250.xlsx')
    print('excel写入结束...')

if __name__ == '__main__':
    main()
