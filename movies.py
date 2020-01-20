import time

from bs4 import BeautifulSoup
from openpyxl import Workbook

from download_page import download_page

DOWNLOAD_URL = 'http://movie.douban.com/top250/'
SLEEP_TIME = 3





def get_li(doc):
    bs = BeautifulSoup(doc, 'html.parser')
    ol = bs.find('ol', class_='grid_view')
    title_list = []
    playable_list = []
    rating_num_list = []
    rating_count_list = []
    inq_list = []
    for i in ol.find_all('li'):
        hd = i.find('div', class_='hd')
        title = hd.find('span', class_='title').get_text()
        title_list.append(title)
        playable = hd.find('span', class_='playable')
        if playable:
            playable_list.append(playable.get_text())
        else:
            playable_list.append('无')
        rating_num = i.find('span', class_='rating_num').get_text()
        rating_num_list.append(rating_num)
        rating_count = i.find('div', class_='star').find_all('span')[-1].get_text()
        rating_count_list.append(rating_count)
        inq = i.find('span', class_='inq')
        if inq:
            inq_list.append(inq.get_text())
        else:
            inq_list.append('无')

    next = bs.find('span', class_='next').find('a')
    if next:
        return title_list, playable_list, rating_num_list, rating_count_list, inq_list, DOWNLOAD_URL + next['href']
    else:
        return title_list, playable_list, rating_num_list, rating_count_list, inq_list, None


def main():
    print('开始获取数据...')
    url = DOWNLOAD_URL
    title_list = []
    playable_list = []
    rating_num_list = []
    rating_count_list = []
    inq_list = []
    while url:
        data = download_page(url)
        title, playable, rating_num, rating_count, inq, url = get_li(data)

        title_list = title_list + title
        playable_list = playable_list + playable
        rating_num_list = rating_num_list + rating_num
        rating_count_list = rating_count_list + rating_count
        inq_list = inq_list + inq
        time.sleep(SLEEP_TIME)
        print('已获取数据%s条' % len(title_list))
    wb = Workbook()
    ws = wb.active
    ws.title = '豆瓣电影top250'
    print('数据获取结束...')
    print('开始写入excel...')
    for (t, p, rn, rc, i) in zip(title_list, playable_list, rating_num_list, rating_count_list, inq_list):
        index = title_list.index(t) + 1
        ws['A%s' % index] = t
        ws['B%s' % index] = rn
        ws['C%s' % index] = rc
        ws['D%s' % index] = p
        ws['E%s' % index] = i

    wb.save(filename='豆瓣电影top250.xlsx')
    print('excel写入结束...')


if __name__ == "__main__":
    main()
